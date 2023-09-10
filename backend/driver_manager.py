import io
import os
import google.auth
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
from google.oauth2 import service_account
from google_auth_oauthlib.flow import InstalledAppFlow
from openpyxl import load_workbook
import json

# Constants
currentPath = os.path.dirname(os.path.realpath(__file__))
folder_driver = '10ZL-GuZJGJW_9FxfgtBjy0g-ibVhJi8E' # folder id of driver minh
# ============================================================

# AUTHENTICATION
credentials = service_account.Credentials.from_service_account_file(
    currentPath + '/service_account.json',
    scopes=['https://www.googleapis.com/auth/drive', 'https://www.googleapis.com/auth/drive.readonly']
)
service = build('drive', 'v3', credentials=credentials)

# ============================================================

def upload_to_driver(folder_driver, filename):
    file_metadata = {'name': filename, 'parents': [folder_driver]}
    media = MediaFileUpload(currentPath + '/' + filename, mimetype='application/pdf')
    file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
    print(f'File ID: {file.get("id")}')
    print('' + file.get('id') + '/view?usp=share_link')
    # {id-file}/view?usp=share_link

def get_files_from_driver(folder_driver):
    ids = []
    results = service.files().list(
        pageSize=10, fields="nextPageToken, files(id, name)", q=f"'{folder_driver}' in parents").execute()
    items = results.get('files', [])
    if not items:
        print('No files found.')
    else:
        print('Files:')
        for item in items:
            print(f"{item['name']} ({item['id']})")
            ids.append(item['id'])
    
    return ids

def download_file_excel_from_driver(file_id):
    request = service.files().export_media(fileId=file_id, mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    fh = io.FileIO('data_sheets/' + file_id + '.xlsx', 'wb')
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()
    
    return 'data_sheets/' + file_id + ".xlsx"

def excel_to_json(file_path, header, min_row, max_row, min_col, max_col):
    workbook = load_workbook(file_path)
    sheet = workbook.active

    data = []
    min_row = min_row if min_row else sheet.min_row
    max_row = max_row if max_row else sheet.max_row
    min_col = min_col if min_col else sheet.min_col
    max_col = max_col if max_col else sheet.max_col
    

    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        row_data = {}
        for idx, cell in enumerate(row):
            row_data[header[idx]] = cell.value
        data.append(row_data)

    return data


def save_to_json(data, output_file):
    non_empty_data = [row for row in data if any(value is not None for value in row.values())]

    with open(output_file, 'w', encoding='utf-8') as json_file:
        json.dump(non_empty_data, json_file, indent=4, ensure_ascii=False)

def convert_json_file_to_array(file_path):
    with open(file_path, encoding='utf-8') as json_file:
        data = json.load(json_file)
    
    return data

# is_sync = False
# if __name__ == '__main__':
    # upload_to_driver(folder_driver, '0317698129.pdf')
    # file_ids = get_files_from_driver(folder_driver)
    # file_sim_xlsx = sheet_id_boi_sim + '.xlsx'
    # if is_sync == True:
    #     file_sim_xlsx = download_file_excel_from_driver(sheet_id_boi_sim)
    # else:
    #     data = excel_to_json(file_sim_xlsx)
    #     save_to_json(data, sheet_id_boi_sim + '.json')
    #     print(len(convert_json_file_to_array(sheet_id_boi_sim + '.json')))
    

# ID của tệp trên Google Drive
# 1Jt2cBzGISX-gp5LfQm7wT3y2yp723ixI/view?usp=drive_link

data = [
    {       
        "so_que": "1",
        "hinh_que": "1Jt2cBzGISX-gp5LfQm7wT3y2yp723ixI",
    },
    {       
        "so_que": "2",
        "hinh_que": "1Rm9jrf8VZ1ljTem3s-xpalqaAl8AguAD",
    },
    {       
        "so_que": "3",
        "hinh_que": "1xfJMXHhKOlqTx4_9e6vo1lHEj4TNRkjd",
    },
    {       
        "so_que": "4",
        "hinh_que": "1KAAcbYRFindSv4RlxOnS7ql1DH9153bt",
    },
    {       
        "so_que": "5",
        "hinh_que": "1QwIJzs4xUntGqiZM2WajQiWC6ZYCzzaF",
    },
    {       
        "so_que": "6",
        "hinh_que": "1AOo6ZHs7lB2FbXXeMypZjYoF0fIOXCd8",
    },
    {       
        "so_que": "7",
        "hinh_que": "1wsfTjp8cbHznuFp_exGhu4XY2BbJosYL",
    },
    {       
        "so_que": "8",
        "hinh_que": "1EFOEPtIOpt1F5HHZFll88l0qwKZrUHEJ",
    },
    {       
        "so_que": "9",
        "hinh_que": "1PQnvcwh7WD-vgvHZUN_Qk7qt1Z6WRSD1",
    },
    {       
        "so_que": "10",
        "hinh_que": "1Jd6l6pQ2onLh8ZElNqoD9j1l9XuFvygk",
    },
    {       
        "so_que": "11",
        "hinh_que": "1JMgmYUR0PWoWOGMDlYz3Cv-k117HBET6",
    },
    {       
        "so_que": "12",
        "hinh_que": "1VBaeeC3GzqmHvdzYvgQCc6-jC1zVF-5r",
    },
    {       
        "so_que": "13",
        "hinh_que": "1Ny47KYssbfxS1spRYNWOCEmt9Qpzg_ay",
    },
    {       
        "so_que": "14",
        "hinh_que": "1nHAzIqR4PnFsZRj_12ZtUuByz2bm_jzm",
    },
    {       
        "so_que": "15",
        "hinh_que": "1fe5aIBg0wio9n5b-H9WbCxiI0hrGfZVZ",
    },
    {       
        "so_que": "16",
        "hinh_que": "1UHhJPG9MhSZVh1YlFusQkKV2xLa0mrCE",
    },
    {       
        "so_que": "17",
        "hinh_que": "1ojF0sPgHqgUhLmsYjr46W9uglPXLiyJF",
    },
    {       
        "so_que": "18",
        "hinh_que": "1v8Q3UF33As9wWMNye-3ruNHQE-0MdsyX",
    },
    {       
        "so_que": "19",
        "hinh_que": "1444XELlur8ueKOqtRxj8YgSgDPAXtKc5",
    },
    {       
        "so_que": "20",
        "hinh_que": "1f0XFZ8qZXBUfy7PzN4VyGJ4cqUnx8QJD",
    },
    {       
        "so_que": "21",
        "hinh_que": "1agKb7dULqPURVhdy__Gm9DtSuo4uu-VP",
    },
    {       
        "so_que": "22",
        "hinh_que": "1DUpsZenYPC63_b5wovkdNf-B6xIx9GoC",
    },
    {       
        "so_que": "23",
        "hinh_que": "1afth11VZyAN6jURof4yA_KcYtUMXZ8Dj",
    },
    {       
        "so_que": "24",
        "hinh_que": "1vN-9P-HTP4PF2N92fYTAD5qSxaBSk5y0",
    },
    {       
        "so_que": "25",
        "hinh_que": "1ULD2IUrobTpMwIAwub_19VWkORQEDsT3",
    },
    {       
        "so_que": "26",
        "hinh_que": "15-G6RS275h0ewIB-LOotjYlMcf3i2ljr",
    },
    {       
        "so_que": "27",
        "hinh_que": "1FMBEyJFXgNX1E7TCfhjoKj7szRFXu1ID",
    },
    {       
        "so_que": "28",
        "hinh_que": "1-FwHXJYTonSKD4ZgOZuraR40oHBYkiLn",
    },
    {       
        "so_que": "29",
        "hinh_que": "1p_5wIEZvLNckB4itTs3OMhmkSt_lFP3w",
    },
    {       
        "so_que": "30",
        "hinh_que": "1iDd9Y76sx5jIXXBMSQXlOMZR96NEKPMp",
    },
    {       
        "so_que": "31",
        "hinh_que": "1uL-yEzBw2K88H4yMgshNFFFHt7FGMZxV",
    },
    {       
        "so_que": "32",
        "hinh_que": "1LeBYt5g-CgZp8oK7UwhJTwGXCUv5xsAF",
    },
    {       
        "so_que": "33",
        "hinh_que": "1pfuzMTxdDxjhqRHrK5cu3jmAsoxWgGaJ",
    },
    {       
        "so_que": "34",
        "hinh_que": "1z1CCM-EbTKsiwHVnDrkAfuOpjqcZ3mOR",
    },
    {       
        "so_que": "35",
        "hinh_que": "1JSbG9PfF8H_S324ENORTYE9A5gFsJbNk",
    },
    {       
        "so_que": "36",
        "hinh_que": "1pWMPwQvYOVu6RaIOAaBiLxImyuA7_faS",
    },
    {       
        "so_que": "37",
        "hinh_que": "1ctVHot7OJthhgoql4jZ9SVVV_tYxO5Vz",
    },
    {       
        "so_que": "38",
        "hinh_que": "197jiiyTU3GbvIGYcLld5MeWI7HANN3qY",
    },
    {       
        "so_que": "39",
        "hinh_que": "1D4OPluKf4afl9gU-4QdmNpfli2mMKDqm",
    },
    {       
        "so_que": "40",
        "hinh_que": "17mLjke69H6kmIPoTrX0zWidNVfnWUO1H",
    },
    {       
        "so_que": "41",
        "hinh_que": "1RQMWiGz-THUwTrbL7-Xo2ggVmY56D9Pw",
    },
    {       
        "so_que": "42",
        "hinh_que": "1tmo9iscGAm7ejS54-iwFSIIdcZkHgtav",
    },
    {       
        "so_que": "43",
        "hinh_que": "1jFBFKud4dUJ53f3pjBqb8PbzUD-EN8wD",
    },
    {       
        "so_que": "44",
        "hinh_que": "1gjnLbTeyUV09c9pWZbUhSVEI-eUv4ZEr",
    },
    {       
        "so_que": "45",
        "hinh_que": "1A-yJj1-boY4LLBFS6qOuOsuXwd16wePi",
    },
    {       
        "so_que": "46",
        "hinh_que": "1gKUgIatjLfPiEs0qP3fSukcI3iDtIClI",
    },
    {       
        "so_que": "47",
        "hinh_que": "13XU_TBy0fyLtnKO7JINWOTegHcCQAA8J",
    },
    {       
        "so_que": "48",
        "hinh_que": "1Ymd3IS2VOWmi3UvpzgSsqlX5BA9EmaZy",
    },
    {       
        "so_que": "49",
        "hinh_que": "1PirOcvT1Hdv9ib3b5vSNiO8NY_ZPbL5P",
    },
    {       
        "so_que": "50",
        "hinh_que": "148aEXsK5yn7ZwGzUuKeznY7xiS2onVeS",
    },
    {       
        "so_que": "51",
        "hinh_que": "1HRnyZ1iBIvtDZ1uvOb6M7vfhKuOoSaAv",
    },
    {       
        "so_que": "52",
        "hinh_que": "1bG84cwQq2_arxQIgTe84Y1VLBlWqIOfv",
    },
    {       
        "so_que": "53",
        "hinh_que": "17dwkXwJL76G7vo3Lk7yPUHZeMWJeSoRk",
    },
    {       
        "so_que": "54",
        "hinh_que": "1nGxMuJkmGx9AG_2mKeXjfEQeIuu0ilqC",
    },
    {       
        "so_que": "55",
        "hinh_que": "1YHW_IcHdIAK4bhnXj0JVgdK2i_ENKbWb",
    },
    {       
        "so_que": "56",
        "hinh_que": "1XVqZrks9uaiW9w2WGOXpMkuxlS0WsjWc",
    },
    {       
        "so_que": "57",
        "hinh_que": "1xlJA8q53Wpi_5kjGr7Kbnk6hZEg0WCf8",
    },
    {       
        "so_que": "58",
        "hinh_que": "1-s40MegLu6Y75DTkHZqD5UIVJA0EHXu2",
    },
    {       
        "so_que": "59",
        "hinh_que": "1StJjRqSs13D60osSgP6p0oArfix2k7br",
    },
    {       
        "so_que": "60",
        "hinh_que": "13V6JHx4sZI5lMAe9JfUG8BIj7Bs9M0LT",
    },
    {       
        "so_que": "61",
        "hinh_que": "1zXOEz-h-vnHRTdSk6sdapk7mD0EYTkgi",
    },
    {       
        "so_que": "62",
        "hinh_que": "1YCKFlN8NNVyH55DTFyJJBRxLlzwbo8s6",
    },
    {       
        "so_que": "63",
        "hinh_que": "1ef6rwPAffrgHnpXXQIclmnm9KOQNDWNr",
    },
    {       
        "so_que": "64",
        "hinh_que": "1oHwjBAGE60xauv6c6d3-7KTZ39GNOXbJ",
    }
]

for item in data:
    local_file_name = f'D:\projects\sim\simhoptuoi\public\common\\templates\site\images\ques' +  '\\' + item['so_que'] + '.jpg'
    file_id = item['hinh_que']
    request = service.files().get_media(fileId=file_id)
    fh = open(local_file_name, 'wb')
    downloader = MediaIoBaseDownload(fh, request)

    done = False
    while not done:
        status, done = downloader.next_chunk()
        print(f"Download {int(status.progress() * 100)}%.")

    print(f"Download completed! File saved as {local_file_name}")