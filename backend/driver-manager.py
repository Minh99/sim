import io
import os
import google.auth
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload

from google.oauth2 import service_account
from google_auth_oauthlib.flow import InstalledAppFlow

currentPath = os.path.dirname(os.path.realpath(__file__))

credentials = service_account.Credentials.from_service_account_file(
    currentPath + '/service_account.json',
    scopes=['https://www.googleapis.com/auth/drive']
)
service = build('drive', 'v3', credentials=credentials)
folder_driver = '10ZL-GuZJGJW_9FxfgtBjy0g-ibVhJi8E'


def upload_to_driver(folder_driver, filename):
    file_metadata = {'name': filename, 'parents': [folder_driver]}
    media = MediaFileUpload(currentPath + '/' + filename, mimetype='application/pdf')
    file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
    print(f'File ID: {file.get("id")}')
    print('https://drive.google.com/file/d/' + file.get('id') + '/view?usp=share_link')
    # https://drive.google.com/file/d/{id-file}/view?usp=share_link

def get_files_from_driver(folder_driver):
    ids = []
    results = service.files().list(
        pageSize=10, fields="nextPageToken, files(id, name)", q=f"'{folder_driver}' in parents").execute()
    items = results.get('files', [])
    if not items:
        print('No files found.')
    else:
        print('Files:')
        for item in items:
            print(f"{item['name']} ({item['id']})")
            ids.append(item['id'])
    
    return ids

def download_file_excel_from_driver(file_id):
    request = service.files().export_media(fileId=file_id, mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    fh = io.FileIO(file_id + '.xlsx', 'wb')
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()
    
    return file_id + ".xlsx"

def read_file_xlsx_by_sheet(filename, sheet_index=0):
    from openpyxl import load_workbook
    workbook = load_workbook(filename=filename)
    sheet_names = workbook.sheetnames
    worksheet = workbook[sheet_names[sheet_index]]
    data = []
    for row in worksheet.values:
        if len(row) >= 0:
            i = 0
            print(len(row[0]))
            item = tuple()
            while True:
                try:
                    if row[i] != None:
                        item += (row[i],)
                        i += 1
                except:
                    break
            data.append(item)
            
    return data
                

if __name__ == '__main__':
    # upload_to_driver(folder_driver, '0317698129.pdf')
    file_ids = get_files_from_driver(folder_driver)
    file_sim_xlsx = download_file_excel_from_driver(file_ids[0])
    print(read_file_xlsx_by_sheet(file_sim_xlsx))